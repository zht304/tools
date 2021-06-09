#! /usr/bin/env python
# -*- coding: utf-8 -*-
# v01 thomas. used for Quectel automotive sw; 2021/5/15
# tested on python 3.8.6
# 从jira里查某个时间段内（不传入参数默认本周）谁（不传入参数默认账号本人）完成了哪些任务(jira状态切换为resolved/closing/sw_resolved/..)

import sys
import jira
import optparse
import getpass

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import worksheet
from openpyxl.styles import Font, colors, Border, Side, Alignment, PatternFill, Color
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from datetime import date, timedelta


def parse_input_args(input_args):
    init_optparse = optparse.OptionParser(usage="usage: %prog [options] arg", description="show help")
    init_optparse.add_option('-s', '--start', help='start date', dest='s_date')
    init_optparse.add_option('-e', '--end', help='end date', dest='e_date')
    init_optparse.add_option('-p', '--project', help='project name in jira', dest='project')
    init_optparse.add_option('-a', '--assignee', help='assignee', dest='assignee')
    init_optparse.add_option('-r', '--reporter', help='reporter', dest='reporter')
    init_optparse.add_option('-l', '--assignee-list', help='file of assignee list', dest='assignee_list')
    init_optparse.add_option('-j', '--jira-account', help='jira account', dest='acct')
    (options, left_args) = init_optparse.parse_args(input_args)

    return options, left_args


class ResolvedSearcher:
    jira_uri = 'https://ticket.quectel.com'

    def __init__(self, account_, passwd_, kwargs):
        self.jira_account = account_
        self.jira_passwd = passwd_
        self.jira_client = jira.JIRA(self.jira_uri, basic_auth=(account_name, passwd))
        self._process_kwargs(kwargs)
        self.kwargs = kwargs

    def _process_kwargs(self, kwargs):
        today = date.today()
        if not kwargs.s_date:
            kwargs.s_date = today - timedelta(days=today.weekday())
        else:
            kwargs.s_date = date.fromisoformat(kwargs.s_date)
        if not kwargs.e_date:
            kwargs.e_date = today + timedelta(days=6 - today.weekday())
        else:
            kwargs.e_date = date.fromisoformat(kwargs.e_date)

    def _generate_jql(self, who):
        my_jql = 'status not in (OPEN,Assigned,SPM_Assigned,"DO",Working,PENDING,Analysing,APPROVING,' \
                 'ST_Open,ST_Reopen,STL_Checked) '
        kwargs = self.kwargs
        if kwargs.project:
            my_jql += 'AND project = %s' % kwargs.project
        if who:
            my_jql += 'AND assignee was "%s" ' % who
        else:
            my_jql += 'AND assignee was currentUser()'
        if kwargs.reporter:
            my_jql += 'AND reporter = "%s" ' % kwargs.reporter

        my_jql += 'AND updatedDate >= %s ' % kwargs.s_date.isoformat()

        return my_jql

    def search_resolved(self, assignee):
        """

        @param assignee: the assignee you want to search
        """
        my_jql = self._generate_jql(assignee)

        search_results = []
        index = 0
        while True:
            sub_list = self.jira_client.search_issues(my_jql, startAt=index, maxResults=500, expand='changelog')
            if not sub_list:
                break
            search_results.extend(sub_list)
            index += 500

        resolved_issues = {}

        for issue in search_results:
            for history in issue.changelog.histories:
                if not history.author.name == assignee:
                    continue

                change_date = date.fromisoformat(history.created[0:10])
                if change_date < self.kwargs.s_date or self.kwargs.e_date < change_date:
                    continue

                for change_content in history.items:
                    if change_content.field == 'status' and change_content.toString \
                            in ('Resolved', 'Closing', 'Done', 'In Review', 'Quectel已完成', 'SW_Resolved'):
                        if issue.key not in resolved_issues:
                            assignee_name = 'unknown'
                            if issue.fields.assignee:
                                assignee_name = issue.fields.assignee.name
                            resolved_issues[issue.key] = {'project': issue.fields.project.key,
                                                          'status': issue.fields.status.name,
                                                          'reporter': issue.fields.reporter.name,
                                                          'assignee': assignee_name,
                                                          'summary': issue.fields.summary}
        return resolved_issues

class JiraResultSaver:
    def __init__(self, file_name_='report.xlsx'):
        self.file_name = file_name_
        self.workbook = Workbook()

    def _set_cell(self, ws, cell, title, border=Border(), fill=None, font=None, alignment=None):
        '''给单元格cell设置值和格式'''
        cell = ws[cell]
        cell.border = border
        if type(title) is str:
            title = ILLEGAL_CHARACTERS_RE.sub(r'', title)
        cell.value = title
        if alignment:
            cell.alignment = alignment

        if font:
            cell.font = font

        if fill:
            cell.fill = fill


    def _insert_row(self, ws, row, start_col, contextlist, border=Border(), fill=None, font=None, alignment=None):
        index = column_index_from_string(start_col)
        for cl in contextlist:
            col = get_column_letter(index)
            self._set_cell(ws, "%s%d" % (col, row), cl, border, fill, font, alignment)
            index += 1


    def _set_sheet_col_width(self, ws):
        # 设置列宽度
        columnWidth = [['A', 25], ['B', 15], ['C', 25], ['D', 30], ['E', 30], ['F', 200]]
        for cw in columnWidth:
            ws.column_dimensions[cw[0]].width = cw[1]


    def write(self, sheet_title, issue_list):
        ###预设格式
        FILL_BLUE = PatternFill(fgColor=Color(indexed=44), fill_type='solid')
        SIDE_THIN = Side(border_style="thin", color="000000")
        BORDER_THIN = Border(top=SIDE_THIN, left=SIDE_THIN, right=SIDE_THIN, bottom=SIDE_THIN)
        ALIG = Alignment(horizontal='general',  # 水平：常规
                         vertical='bottom',  # 垂直：底部对齐
                         text_rotation=0,  # 文本方向：0度
                         wrap_text=True,  # 自动换行
                         shrink_to_fit=False,  # 缩小字体填充
                         indent=0  # 缩进0
                         )

        ws = self.workbook.create_sheet(sheet_title, 0)
        header = ['Issue', 'Status', 'project', 'reporter', 'assignee', 'summary']
        self._insert_row(ws, 1, 'A', header, fill=FILL_BLUE, border=BORDER_THIN, alignment=ALIG)
        self._set_sheet_col_width(ws)

        row = 2
        for issue_key, issue_data in issue_list.items():
            line = [issue_key, issue_data['status'], issue_data['project'], issue_data['reporter'],
                    issue_data['assignee'], issue_data['summary']]
            self._insert_row(ws, row, 'A', line, border=BORDER_THIN, alignment=ALIG)
            ws.row_dimensions[row].height = 30
            row += 1

    def save(self):
        self.workbook.save(self.file_name)


def read_persons_from_file(persons, list_file):
    with open(list_file, 'r') as f:
        for line in f.readlines():
            name = line.strip()
            if name:
                persons.append(line.strip())


if __name__ == "__main__":
    (opts, args) = parse_input_args(sys.argv[1:])

    if not opts.acct:
        account_name = input('Jira Account: ')
    else:
        account_name = opts.acct
        print('Jira account: %s' % account_name)
    passwd = getpass.getpass('Jira Password: ')

    team = []
    if not opts.assignee_list:
        if not opts.assignee:
            opts.assignee = account_name
        team.append(opts.assignee)
    else:
        read_persons_from_file(team, opts.assignee_list)

    searcher = ResolvedSearcher(account_name, passwd, opts)
    excel = JiraResultSaver('report.xlsx')

    for engineer in team:
        print('searching issues of %s...' % engineer)
        issues = searcher.search_resolved('%s@quectel.com' % engineer)
        excel.write(engineer, issues)
        # save at once to reduce RAM occupy
        excel.save()

    print('Done')
